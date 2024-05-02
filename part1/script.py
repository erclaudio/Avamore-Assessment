from openpyxl import load_workbook
from datetime import datetime
from time import sleep
import sys
import xlwings
import os
def load_excel():
    return load_workbook('part1/ARM Template_Simplified.xlsx')

def save_and_close(workbook):
    if workbook:
        try:
            workbook.save('part1/ARM Template_Simplified.xlsx')
            workbook.close()
            excel_app = xlwings.App(visible=False)
            excel_book = excel_app.books.open('part1/ARM Template_Simplified.xlsx')
            excel_book.save()
            excel_book.close()
            excel_app.quit()
        except Exception as e:
            print("Error saving and closing workbook:", e)
        

def facility_A_value(workbook):
    inputs_sheet = workbook['Inputs']
    inputs_sheet['C15'] = float(input("Enter Facility A Value: "))
    save_and_close(workbook)

def change_interest(workbook):
    inputs_sheet = workbook['Inputs']
    interest_type = input("Press 1 to enter monthly Interest.\nPress 2 to enter yearly interest.\n:")
    interest_rate = float(input("Enter Monthly Interest rate: ")) / 100 if interest_type == '1' else float(input("Enter Yearly Interest rate: ")) / 12 / 100
    inputs_sheet['C22'] = interest_rate
    save_and_close(workbook)

def change_default_date(workbook, column):
    inputs_sheet = workbook['Inputs']
    date_input = input(f"Enter Default {column} Date (YYYY-MM-DD): ")
    inputs_sheet[f'C28' if column == 'Start' else 'C29'] = datetime.strptime(date_input, "%Y-%m-%d")
    save_and_close(workbook)

def view_details():
    workbook = load_workbook('part1/ARM Template_Simplified.xlsx', data_only=True)
    inputs_sheet = workbook['Inputs']
    facility_cell_value = inputs_sheet['C15'].value
    interest_cell_value = round(inputs_sheet['C22'].value * 100, 3)
    def_start_cell_value = inputs_sheet['C28'].value
    def_end_cell_value = inputs_sheet['C29'].value
    print(f'{"Facility A Value:":22} {facility_cell_value}\n'
          f'{"Interest Rate:":22} {interest_cell_value}\n'
          f'{"Default Period Start:":22} {def_start_cell_value}\n'
          f'{"Default Period End:":22} {def_end_cell_value}')
    workbook.close()

def view_interest():
    try:
        excel_app = xlwings.App(visible=False)
        excel_book = excel_app.books.open('part1/ARM Template_Simplified.xlsx')
        excel_book.save()
        excel_book.close()
        excel_app.quit()

        data_workbook = load_workbook('part1/ARM Template_Simplified.xlsx', data_only=True)
        data_calculations_sheet = data_workbook['Calculations']
        total_interest = round(data_calculations_sheet['B3'].value, 2)
        print(f'{"Total Interest Due:":22} {total_interest}')
        data_workbook.close()
    except Exception as e:
        print("An error occurred:", e)


def exit_menu():
    print("Changes Saved, Exiting...")
    sys.exit()

def main():
    while True:
        print("""
        MENU
        --------------------
        1. Change Facility A value
        2. Change Interest Rate
        3. Change Beginning of Default Period
        4. Change End of Default Period
        5. View Inputs & Total Interest Due
        6. Exit and Save Changes
        """)
        choice = input("Enter your selection: ")
        if choice == '1':
            facility_A_value(load_excel())
        elif choice == '2':
            change_interest(load_excel())
        elif choice == '3':
            change_default_date(load_excel(), "Start")
        elif choice == '4':
            change_default_date(load_excel(), "End")
        elif choice == '5':
            view_details()
            view_interest()
        elif choice == '6':
            exit_menu()
        else:
            print("Please enter a valid selection.")

if __name__=="__main__":
    main()
